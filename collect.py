import feedparser, schedule, time, os, pytz
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 엑셀 파일명 정의
filename = "google_trends_data.xlsx"


def get_first_news_item(entry):
    """
    RSS feed entry에서 첫 번째 뉴스 아이템을 추출하여 반환한다.
    """
    if 'ht_news_item' in entry and isinstance(entry['ht_news_item'], list) and len(entry['ht_news_item']) > 0:
        first_news_item = entry['ht_news_item'][0]
        return {
            'News Item Title': first_news_item.get('ht_news_item_title', ''),
            'News Item URL': first_news_item.get('ht_news_item_url', ''),
            'News Item Source': first_news_item.get('ht_news_item_source', '')
        }
    else:
        return {
            'News Item Title': '',
            'News Item URL': '',
            'News Item Source': ''
        }


def fetch_and_save_yesterdays_data():
    feed_url = 'https://trends.google.co.kr/trends/trendingsearches/daily/rss?geo=KR'
    feed = feedparser.parse(feed_url)

    # 한국 시간대 설정
    KST = pytz.timezone('Asia/Seoul')


    # 어제 게시된 포스트만 수집
    entries = []
    for entry in feed.entries:
        # pubDate을 한국 시간대로 변환
        pubDate = datetime(*entry.published_parsed[:6], tzinfo=pytz.utc).astimezone(KST)
        yesterday = datetime.now(KST) - timedelta(days=1)
        if pubDate.date() == yesterday.date():
            news_item = get_first_news_item(entry)
            entries.append({
                'Title': entry.title,
                'Approx Traffic': entry.get('ht_approx_traffic'),
                'PubDate': pubDate.strftime('%Y-%m-%d %H:%M:%S'),
                'News Item Title': entry.get('ht_news_item_title', ''),
                'News Item URL': entry.get('ht_news_item_url', ''),
                'News Item Source': entry.get('ht_news_item_source', '')
            })

    if entries:
        new_data = pd.DataFrame(entries)

        # 날짜 헤더
        date_header = pd.DataFrame(
            {'Title': [yesterday.strftime('%Y-%m-%d')], 'Approx Traffic': [""], 'Link': [""], 'PubDate': [""],
             'News Item Title': [""], 'News Item URL': [""], 'News Item Source': [""]})

        # 날짜 헤더와 새 데이터 결합
        data_to_save = pd.concat([date_header, new_data], ignore_index=True)

    # 파일이 이미 존재하는 경우 기존 파일에 데이터 추가
    if os.path.exists(filename):
        book = load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = writer.sheets['Sheet1'].max_row
            data_to_save.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=False)
    else:
        # 파일이 없는 경우 새 파일 생성
        data_to_save.to_excel(filename, index=False)


schedule.every().day.at("00:01").do(fetch_and_save_yesterdays_data)

while True:
    schedule.run_pending()
    time.sleep(60)  # 60초 간격으로 체크
